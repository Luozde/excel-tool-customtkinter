from collections import defaultdict

import openpyxl

from core.Product import Detail, Product, Key2
from core.Sku import Sku

# 用于将Excel文件转换为Sku对象列表
class DataTransformer:
    def __init__(self):
        # 可以在初始化方法中设置必要的属性或初始化变量
        pass

    # 用于将Excel文件转换为Sku对象列表
    def process_excel_file(app, file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        skuItemList = []
        headers = [str(cell.value).strip() for cell in ws[1]]
        try:
            sku_index = headers.index('SKU')
            quantity_index = headers.index('产品数量')
            image_index = headers.index('图片')
            order_no_index = headers.index('订单号')
        except Exception as e:
            app.print_logs("订单文件校验:\n", "end")
            app.print_logs("可能未找到【SKU】,【产品数量】,【图片】或【订单号】列，具体请看《异常信息》\n", "end")
            app.print_logs("异常信息：" + str({e}) + "\n", "end")
            return skuItemList

        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            try:
                skuStr = row[sku_index]
                try:
                    quantity = int(row[quantity_index])
                except Exception:
                    app.print_logs(f"第{row_index + 1}行-订单号【{row[order_no_index]}】: 产品数量不规范，未被统计入内，请人工核对\n","end")
                    continue
                image = row[image_index]

                skuItem = Sku.create_sku(skuStr)
                skuItem.quantity = quantity
                skuItem.image = image

                for image in ws._images:
                    if image.anchor._from.row == row_index:
                        image.format = 'png'
                        skuItem.imageFile = image
                        break
                skuItemList.append(skuItem)
            except (Exception):
                app.print_logs(f"第{row_index+1}行-订单号【{row[order_no_index]}】: SKU【{row[sku_index]}】不规范，未被统计入内，请人工核对\n", "end")
                continue

        app.print_logs("读入订单总数：" + str(ws.max_row - 1) + "\n", "end")
        app.print_logs("处理有效订单数：" + str(len(skuItemList)) + "\n", "end")
        app.print_logs("未参与统计订单数：" + str(ws.max_row-1 - len(skuItemList)) + "\n", "end")

        return skuItemList

    # 用于将Sku对象列表转换为Product对象列表
    def summarize_sku_items(app, skuItemList):
        # product_dict = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
        product_dict = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))


        for skuItem in skuItemList:
            productNo = skuItem.productNo
            color = skuItem.color
            index = skuItem.index
            text = skuItem.text
            size = skuItem.size
            quantity = skuItem.quantity
            key1 = skuItem.key1
            key2 = skuItem.key2

            product_dict[key1][key2][color]["image"] = skuItem.image
            product_dict[key1][key2][color]["imageFile"] = skuItem.imageFile
            product_dict[key1][key2][color][size] += quantity
            product_dict[key1][key2][color]["text"] = text
            product_dict[key1][key2][color]["sex"] = skuItem.sex
            product_dict[key1][key2][color]["color"] = color

        key1_list = []
        for key1, key1_dict in product_dict.items():
            key2_list = []
            for key2, key2_dict in key1_dict.items():
                details = []
                for color, detail_dict in key2_dict.items():
                    detail = Detail(
                        image=detail_dict["image"],
                        imageFile=detail_dict["imageFile"],
                        color=detail_dict["color"],
                        text=detail_dict["text"],
                        sex=detail_dict["sex"],
                        s=detail_dict["S"],
                        m=detail_dict["M"],
                        l=detail_dict["L"],
                        l1=detail_dict["XL"],
                        l2=detail_dict["2XL"],
                        l3=detail_dict["3XL"],
                        l4=detail_dict["4XL"],
                        l5=detail_dict["5XL"]
                    )
                    details.append(detail)
                key2_item = Key2(key2, details)
                key2_list.append(key2_item)

                # 对details进行排序
                # details.sort(key=lambda x: x.color)

            product = Product(key2_list, key1)
            key1_list.append(product)

            # 对key2_list进行排序
            key1_list.sort(key=lambda x: x.key1)

        return key1_list

    from collections import defaultdict

    def summarize_sku_items1(app, skuItemList):
        product_dict = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))

        for skuItem in skuItemList:
            product_dict[skuItem.productNo][skuItem.index][skuItem.color].update({
                "image": skuItem.image,
                "imageFile": skuItem.imageFile,
                skuItem.size: product_dict[skuItem.productNo][skuItem.index][skuItem.color][
                                  skuItem.size] + skuItem.quantity,
                "text": skuItem.text,
                "sex": skuItem.sex,
                "color": skuItem.color
            })

        productList = []
        for key1, index_dict in product_dict.items():
            details = []
            for key2, color_dict in index_dict.items():
                for color, size_dict in color_dict.items():
                    detail = Detail(
                        image=size_dict["image"],
                        imageFile=size_dict["imageFile"],
                        color=size_dict["color"],
                        text=size_dict["text"],
                        sex=size_dict["sex"],
                        **{size: size_dict[size] for size in ["S", "M", "L", "XL", "2XL", "3XL", "4XL", "5XL"]}
                    )
                    details.append(detail)

            # 对details进行排序
            details.sort(key=lambda x: x.color)

            product = Product(key1, details)
            productList.append(product)

        return productList


