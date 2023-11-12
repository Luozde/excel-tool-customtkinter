import json

import openpyxl
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU

# from core import ExcelExporter
class ExcelExporter:
    @staticmethod
    def export_to_excel(app, product_list, file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        # 设置表头
        headers = ['性别款式', '商品号', '图像', '颜色', 'S', 'M', 'L', 'XL', '2XL', '3XL', '合计']
        header_font = Font(bold=True, size=14)
        header_alignment = Alignment(horizontal='center', vertical='center')
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment

        # 导出数据
        row_index = 2
        data_font = Font(size=14)
        data_font_bold = Font(size=14, color='FFFF0000', bold=True)
        data_alignment = Alignment(vertical='center')
        noIndex = 0
        for product in product_list:
            parts = product.key1.split("-")
            sex = '通款'
            if len(parts) == 2:
                sex = parts[1]

            key2_item = product.key2_list[0]
            product_no = key2_item.key2
            key2_list = product.key2_list
            merged_cell = ws.cell(row=row_index, column=2, value=product_no)
            merged_cell.alignment = Alignment(horizontal='center', vertical='center')
            merged_cell = ws.cell(row=row_index, column=1, value=sex)
            merged_cell.alignment = Alignment(horizontal='center', vertical='center')
            key1_count = 0
            key1_row_index = row_index
            for key2 in key2_list:
                details = key2.details
                product_no = key2.key2
                #if len(details) > 1:
                # 合并productNo单元格
                ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index + len(details) - 1, end_column=2)
                merged_cell = ws.cell(row=row_index, column=2, value=product_no)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                key1_count += len(details)

                for detail in details:
                    # ws.cell(row=row_index, column=3, value=detail.image).font = data_font
                    ws.cell(row=row_index, column=4, value=detail.color).font = data_font
                    ws.cell(row=row_index, column=5, value=detail.s).font = data_font
                    ws.cell(row=row_index, column=6, value=detail.m).font = data_font
                    ws.cell(row=row_index, column=7, value=detail.l).font = data_font
                    ws.cell(row=row_index, column=8, value=detail.l1).font = data_font
                    ws.cell(row=row_index, column=9, value=detail.l2).font = data_font
                    ws.cell(row=row_index, column=10, value=detail.l3).font = data_font
                    # 计算合计值
                    total_formula = f"=SUM(E{row_index}:J{row_index})"
                    ws.cell(row=row_index, column=11, value=total_formula).font = data_font

                    img = detail.imageFile
                    # if img is not None:
                    #     ExcelExporter.insert_image(ws,'c', row_index,1, img)
                    if img is not None:
                        img.width = 100
                        img.height = 100
                        ExcelExporter.offset_img(img, 2, row_index-1)
                        ws.add_image(img)
                    row_index += 1

            # 合并key1单元格
            if key1_count > 1:
                ws.merge_cells(start_row=key1_row_index, start_column=1, end_row=key1_row_index + key1_count - 1, end_column=1)
                merged_cell = ws.cell(row=row_index, column=1, value=sex)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 设置列宽
        column_widths = [20, 20, 20, 20, 10, 10, 10, 10, 10, 10, 10]
        for col_num, width in enumerate(column_widths, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = width

        # 设置行高
        ws.row_dimensions[1].height = 30
        for row in ws.iter_rows(min_row=2, min_col=1, max_row=row_index - 1, max_col=11):
            for cell in row:
                cell.alignment = data_alignment
                cell.font = data_font
                size_start = 5
                total_index = size_start + 6
                try:
                    if cell.column >= size_start and cell.column < total_index and cell.value is not None and int(cell.value) > 0:
                        cell.font = data_font_bold
                    if cell.column == total_index:
                        cell.font = Font(size=16, bold=True)
                except:
                    print(cell.value)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.row_dimensions[cell.row].height = 80

        ws.freeze_panes = 'A2'
        wb.save(file_path)

    @staticmethod
    def export_to_json(product_list, file_path):
        data = []
        for product in product_list:
            product_dict = {
                'productNo': product.productNo,
                'details': []
            }
            for detail in product.details:
                detail_dict = {
                    'image': detail.image,
                    'color': detail.color,
                    'S': detail.s,
                    'M': detail.m,
                    'L': detail.l,
                    'XL': detail.l1,
                    '2XL': detail.l2,
                    '3XL': detail.l3
                }
                product_dict['details'].append(detail_dict)
            data.append(product_dict)

        with open(file_path, 'w') as json_file:
            json.dump(data, json_file, indent=4)

    @staticmethod
    def offset_img(img, x, y):
        """精确设置图片位置，偏移量以万为单位进行微调吧，具体计算公式太麻烦了
        row column 的索引都是从0开始的，我这里要把图片插入到单元格B10
        """
        p2e = pixels_to_EMU
        h, w = img.height, img.width
        size = XDRPositiveSize2D(p2e(w), p2e(h))
        marker = AnchorMarker(col=2, colOff=300000, row=y, rowOff=30000)
        img.anchor = OneCellAnchor(_from=marker, ext=size)

